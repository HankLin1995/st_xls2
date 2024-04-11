import streamlit as st
from data_stream import stream_data  # 導入資料流模組
import math
import sympy as sp
import numpy as np
import openpyxl
import os

def calculate_active_pressure_coefficient(phi):
    Ka = (1 - math.sin(math.radians(phi))) / (1 + math.sin(math.radians(phi)))
    return Ka

def calculate_passive_pressure_coefficient(phi):
    Kp = (1 + math.sin(math.radians(phi))) / (1 - math.sin(math.radians(phi)))
    return Kp

def resolve_d0(Ka,Kp,H,FS,r):

    d0 = sp.Symbol('d0')

    Mr = (1/2) * r * (H + d0)**2 * Ka * ((H + d0) / 3)
    Ml = (1/2) * r * d0**2 * Kp * (d0 / 3)        # 彎矩平衡方程
    equation = sp.Eq(Ml, FS*Mr)

    # 解方程，得到d0
    d0_solutions = sp.solve(equation, d0)
    # 筛选出正的解
    positive_solutions = [sol.evalf() for sol in d0_solutions if sol.is_real and sol.evalf() > 0]

    for sol in positive_solutions:
        d0_sol=float(sol)
    return d0_sol
    
def resolve_D(d0_sol, r, H,R, Ka, Kp):
    i_value = None
    S_test_value = None
    for i in np.arange(1.1, 1.3, 0.01):
        D_test = i * d0_sol
        P_Al = (1/2) * (r * d0_sol * Ka + r * D_test * Ka) * (D_test - d0_sol)
        P_pl = (1/2) * (r * (H + d0_sol) * Kp + r * (H + D_test) * Kp) * (D_test - d0_sol)
        S_test = P_pl - P_Al
        if S_test > R:
            i_value = i
            S_test_value = S_test
            break
    return i_value, S_test_value
    

def SP():
    st.header(":memo: 懸臂式板樁")
    # 創建一個表單
    with st.form("input_form"):
        st.write("**已知條件：**")
        st.write("")
        col1, col2 = st.columns(2)

        with col1:

            H = st.number_input("開挖深度[H]", value=5.0,help="單位=公尺")
            r = st.number_input("土壤單位重", value=1.8,help="計算上不影響結果")
            phi = st.number_input("土壤安息角", value=30,help="常見為30~33")
            FS = st.number_input("安全係數", value=1.5,help="規範建議>=1.5")

        with col2:
            # st.image('st_xls2/photos/St.jpg', caption='Fig1.懸臂式板樁範例')
            st.image('photos/St.jpg', caption='Fig1.懸臂式板樁範例')
        # 表單提交按鈕
        submit_button = st.form_submit_button("計算",type="primary")

    
    if submit_button:

        #計算過程

        Ka = calculate_active_pressure_coefficient(phi)
        Kp = calculate_passive_pressure_coefficient(phi)

        d0_sol=resolve_d0(Ka,Kp,H,FS,r)
        
        # 代入 d0 的解到 P_A 和 P_p 的表达式中
        P_A_value = (1/2) *r * (H + d0_sol)**2 * Ka
        P_p_value = (1/2) * r * (d0_sol)**2 * Kp

        # 计算 R
        R_value = P_p_value - P_A_value
        i_value, S_test_value = resolve_D(d0_sol, r, H,R_value, Ka, Kp)

        D=i_value*d0_sol

        st.toast(":heavy_check_mark: 建議採用SP為"+str(math.ceil(H + D)) +"M")

        #----說明開始----

        with open("./md/SP.md", "r", encoding="utf-8") as file:
            markdown_text = file.read()

        markdown_text = markdown_text.format(
            phi=phi,
            Ka=Ka,
            Kp=Kp,
            d0_sol=d0_sol,
            R_value=R_value,
            i_value=round(i_value, 2),
            S_test_value=S_test_value, 
            D=i_value * d0_sol,
            H_D=H + D,
            H_D_ceil=math.ceil(H + D)
        )

        st.toast(":speech_balloon: 開始列出計算流程")

        # 在 Streamlit 中呈现 Markdown 内容
        st.markdown(markdown_text)

        st.toast(":file_folder: 文件下載已準備完成!")

        #----Excel報表內容填寫----

        workbook = openpyxl.load_workbook('./template/SP.xlsx')

        sheet = workbook.active

        # 将数字写入指定单元格，例如将数字 123 写入第一行第一列的单元格
        sheet.cell(row=1, column=1).value = st.session_state['company']
        sheet.cell(row=3, column=3).value = st.session_state['conName']
        sheet.cell(row=4, column=3).value = st.session_state['conLoc']
        
        sheet.cell(row=12, column=3).value = H
        sheet.cell(row=13, column=3).value = phi
        sheet.cell(row=14, column=3).value = FS
        sheet.cell(row=16, column=5).value = Ka
        sheet.cell(row=19, column=5).value = Kp

        sheet.cell(row=36, column=3).value = d0_sol
        sheet.cell(row=40, column=3).value = R_value
        sheet.cell(row=51, column=2).value = i_value
        sheet.cell(row=51, column=4).value = S_test_value
        sheet.cell(row=52,column=4).value=D
        sheet.cell(row=56, column=4).value = H+D
        sheet.cell(row=58, column=5).value = math.ceil(H + D)

        output_file = 'example.xlsx'
        workbook.save(output_file)

        # 提供下载链接给用户

        with open(output_file, 'rb') as f:
            bytes_data = f.read()
        st.download_button(label='計算成果下載', data=bytes_data, file_name=output_file,type='primary')

        os.remove(output_file)
            
    st.session_state.current_page = 'SP'
