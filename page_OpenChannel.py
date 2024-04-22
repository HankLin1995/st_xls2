import streamlit as st
from data_stream import stream_data  # 導入資料流模組
import math
import sympy as sp
import pandas as pd
import openpyxl
import os

def Uchannel_FixB():
    st.subheader(":bookmark_tabs: U型溝計算(固定寬度)")
    # 創建一個表單
    with st.form("input_form"):
        st.write("**已知條件：**")

        col1, col2 = st.columns([1,2])

        with col1:
           
            q = st.number_input("計畫流量q (cms)", value=18.0)
            n = st.number_input("曼寧係數n", value=0.0180, format="%.3f")
            s = st.number_input("渠道縱坡 s=1/", value=1000)
            b = st.number_input("渠寬b (m)", value=6.0)
            fill_angle = st.number_input("填角F (m)", value=0.0)
            H = st.number_input("渠牆高度H (m)", value=3.0)
        with col2:

            st.image('photos/images.jpg', caption='Fig1.U型溝照片範例')
            is_download=st.toggle("檢核未過還是要下載")

        # 表單提交按鈕
        submit_button = st.form_submit_button("計算",type="primary")


    if submit_button:

        #---計算開始---

        ## 計算y
        y = sp.Symbol('y')

        P = (b-2*fill_angle)+2*math.sqrt(2)*fill_angle + 2*(y-fill_angle)
        A = b * y-fill_angle**2
        R = A / P
        V=(1 / n) * R** (2/3) * math.sqrt(1/s)
        Q = A * V

        solution = sp.solve(Q - q, y)
        y_sol=float(solution[0])
        # st.write("y=" ,y_sol,"m")

        ## 帶入 y,計算Fr,V
        y=y_sol

        P = (b-2*fill_angle)+2*math.sqrt(2)*fill_angle + 2*(y-fill_angle)
        A = b * y-fill_angle**2
        R = A / P
        V=(1 / n) * R** (2/3) * math.sqrt(1/s)
        Q = A * V
        T=b
        Fr=V/(9.81*A/T)**0.5
        hv=V**2/(2*9.81)

        # st.write("V =" ,V,"m/s")
        # st.write("Fr =" ,Fr)
        # st.write("Q=",Q,"cms")
        # st.write("hv=",hv,"m")

        ## 計算yc

        yc=sp.Symbol('yc')
        Tc=b
        Ac=b*yc

        eq=q**2*Tc-9.81*(Ac**3)

        solution2 = sp.solve(eq,yc)
        yc_sol=float(solution2[0])
        # st.write("yc=" ,yc_sol,"m")

        Vc=Q/(yc_sol*b)
        # st.write("Vc =" ,Vc,"m/s")

        #---過程描述---#

        with open("./md/U.md", "r", encoding="utf-8") as file:
            markdown_text = file.read()

        markdown_text = markdown_text.format(
            q=q,
            n=n,
            B=b,
            F=fill_angle,
            s=s,
            y=round(y,3),
            Q=round(Q,3),
            V=round(V,3),
            Fr=round(Fr,3),
            yc=round(yc_sol,3),
            Vc=round(Vc,3),
            Wall=1

        )

        st.markdown(markdown_text)

        #---計算檢核---#

        st.write("**檢核過程：**")
        err_state=False
        if V>0.7:
            st.write(":white_check_mark:最小容許流速 " )
        else:
            st.error('	:x:最小容許流速請再檢討')
            st.write("設計流速",V,"<最小容許流速",0.7,"m/s")
            err_state=True

        if V<3:
            st.write(":white_check_mark:最大容許流速(農工手冊) " )
        elif V<4:
            st.write(":white_check_mark:最大容許流速(合訂本)) " )
        else:
            st.error('	:x:最大容許流速請再檢討')
            st.write("設計流速",V,">最大容許流速(合訂本)",4,"m/s")
            err_state=True

        if V<0.66*Vc:
            st.write(":white_check_mark:避免臨界流速(農工手冊)" )
        elif V<0.8*Vc:
            st.write(":white_check_mark:避免臨界流速(合訂本) " )
        else:
            st.error('	:x:臨界流速請再檢討')
            st.write("設計流速",V,">0.8*臨界流速(合訂本)",0.8*Vc,"m/s")
            err_state=True

        Fb1=max(1/3*y,0.15)

        Fb2=0.07*y+hv+0.15

        df = pd.DataFrame({
            '合訂本': [Fb1],
            '農工手冊': [Fb2],
        })

        Fb=max(Fb1,Fb2)

        y_ans=round(y+Fb,3)

        st.write(":white_check_mark: 出水高度選定")
        st.dataframe(df,hide_index=True)
        st.write("出水高度採**最大值**為" ,round(Fb,3) ," M")

        if H>y_ans:
            st.write(":white_check_mark:牆高是否足夠")
        else:
            st.error('	:x:牆高請再檢討')
            st.write("需求牆高",y_ans,"<設計牆高",H)
            err_state=True

        if err_state==False:
            st.markdown("---")
            st.subheader("六、計算成果")

            data = pd.DataFrame({
                'Q(cms）': [Q],
                'B（m）': [b],
                'n': [n],
                '1/s': [1/s],
                'd（m）': [y_ans],
                'A（m2）': [A],
                'P（m）': [P],
                'R': [R],
                'V（m/s）': [V],
                'hv（m）': [hv]
            })

            st.dataframe(data,hide_index=True)

            data = pd.DataFrame({
                '設計流量': [Q],
                '渠道寬度': [b],
                '曼寧係數': [n],
                '坡度': [1/s],
                '設計水深': [y_ans],
                '面積': [A],
                '潤周': [P],
                '水力半徑': [R],
                '設計流速': [V],
                '速度水頭': [hv]
            })

            st.dataframe(data,hide_index=True)

             #----Excel報表內容填寫----

        if (err_state==True and is_download==True) or err_state==False:

            workbook = openpyxl.load_workbook('./template/U.xlsx')

            sheet = workbook.active

            # 将数字写入指定单元格，例如将数字 123 写入第一行第一列的单元格
            sheet.cell(row=1, column=1).value = st.session_state['company']
            sheet.cell(row=3, column=3).value = st.session_state['conName']
            sheet.cell(row=4, column=3).value = st.session_state['conLoc']
            
            sheet.cell(row=12, column=3).value = q
            sheet.cell(row=13, column=3).value = n
            sheet.cell(row=14, column=3).value = b
            sheet.cell(row=15, column=3).value = fill_angle
            sheet.cell(row=16, column=3).value = 1/s

            sheet.cell(row=21, column=3).value = y
            sheet.cell(row=22, column=3).value = Fb
            sheet.cell(row=24, column=8).value = P
            sheet.cell(row=25, column=5).value = A
            sheet.cell(row=26,column=5).value=R
            sheet.cell(row=27, column=6).value = V
            sheet.cell(row=28, column=5).value = Q

            sheet.cell(row=29, column=3).value = hv
            sheet.cell(row=30, column=3).value = Fr
            sheet.cell(row=32, column=3).value = Vc
            sheet.cell(row=33, column=3).value = yc_sol
            sheet.cell(row=67, column=4).value = H

            output_file = 'example.xlsx'
            workbook.save(output_file)
            
            with open(output_file, 'rb') as f:
                bytes_data = f.read()
            st.download_button(label='計算成果下載', data=bytes_data, file_name=output_file,type='primary')

            os.remove(output_file)

    st.session_state.current_page = 'Uchannel_FixB'


def Uchannel_FixY():
    st.subheader(":bookmark_tabs: U型溝計算(固定水深)")
    # 創建一個表單
    with st.form("input_form"):
        st.write("**已知條件：**")

        col1, col2 = st.columns([1,2])

        with col1:
           
            q = st.number_input("計畫流量q (cms)", value=18.0)
            n = st.number_input("曼寧係數n", value=0.0180, format="%.3f")
            s = st.number_input("渠道縱坡 s=1/", value=1000)
            y = st.number_input("水深y (m)", value=1.5)
            fill_angle = st.number_input("填角F (m)", value=0.0)
        with col2:

            st.image('photos/images.jpg', caption='Fig1.U型溝照片範例')
            is_download=st.toggle("檢核未過還是要下載")

        # 表單提交按鈕
        submit_button = st.form_submit_button("計算",type="primary")


    if submit_button:

        #---計算開始---

        ## 計算b
        b=sp.Symbol('b')

        P = (b-2*fill_angle)+2*math.sqrt(2)*fill_angle + 2*(y-fill_angle)
        A = b * y-fill_angle**2
        R = A / P
        V=(1 / n) * R** (2/3) * math.sqrt(1/s)
        Q = A * V

        solution = sp.solve(Q - q, b)
        b_sol=float(solution[0])
        # st.write("b=" ,b_sol,"m")

        ## 帶入 b,計算Fr,V
        b=b_sol

        P = (b-2*fill_angle)+2*math.sqrt(2)*fill_angle + 2*(y-fill_angle)
        A = b * y-fill_angle**2
        R = A / P
        V=(1 / n) * R** (2/3) * math.sqrt(1/s)
        Q = A * V
        T=b
        Fr=V/(9.81*A/T)**0.5
        hv=V**2/(2*9.81)

        # st.write("V =" ,V,"m/s")
        # st.write("Fr =" ,Fr)
        # st.write("Q=",Q,"cms")
        # st.write("hv=",hv,"m")

        ## 計算yc

        yc=sp.Symbol('yc')
        Tc=b
        Ac=b*yc

        eq=q**2*Tc-9.81*(Ac**3)

        solution2 = sp.solve(eq,yc)
        yc_sol=float(solution2[0])
        # st.write("yc=" ,yc_sol,"m")

        Vc=Q/(yc_sol*b)
        # st.write("Vc =" ,Vc,"m/s")

        #---過程描述---#

        with open("./md/U2.md", "r", encoding="utf-8") as file:
            markdown_text = file.read()

        markdown_text = markdown_text.format(
            q=q,
            n=n,
            B=round(b,3),
            F=fill_angle,
            s=s,
            y=round(y,3),
            Q=round(Q,3),
            V=round(V,3),
            Fr=round(Fr,3),
            yc=round(yc_sol,3),
            Vc=round(Vc,3),
            Wall=1

        )

        st.markdown(markdown_text)

        #---計算檢核---#

        st.write("**檢核過程：**")
        err_state=False
        if V>0.7:
            st.write(":white_check_mark:最小容許流速 " )
        else:
            st.error('	:x:最小容許流速請再檢討')
            st.write("設計流速",V,"<最小容許流速",0.7,"m/s")
            err_state=True

        if V<3:
            st.write(":white_check_mark:最大容許流速(農工手冊) " )
        elif V<4:
            st.write(":white_check_mark:最大容許流速(合訂本)) " )
        else:
            st.error('	:x:最大容許流速請再檢討')
            st.write("設計流速",V,">最大容許流速(合訂本)",4,"m/s")
            err_state=True

        if V<0.66*Vc:
            st.write(":white_check_mark:避免臨界流速(農工手冊)" )
        elif V<0.8*Vc:
            st.write(":white_check_mark:避免臨界流速(合訂本) " )
        else:
            st.error('	:x:臨界流速請再檢討')
            st.write("設計流速",V,">0.8*臨界流速(合訂本)",0.8*Vc,"m/s")
            err_state=True

        Fb1=max(1/3*y,0.15)

        Fb2=0.07*y+hv+0.15

        df = pd.DataFrame({
            '合訂本': [Fb1],
            '農工手冊': [Fb2],
        })

        Fb=max(Fb1,Fb2)

        y_ans=round(y+Fb,3)

        st.write(":white_check_mark: 出水高度選定")
        st.dataframe(df,hide_index=True)
        st.write("出水高度採**最大值**為" ,round(Fb,3) ," M")

        # if B_target>b:
        #     st.write(":white_check_mark:渠寬是否足夠")
        # else:
        #     st.error('	:x:渠寬請再檢討')
        #     st.write("需求牆高",y_ans,"<設計牆高",H)
        #     err_state=True

        if err_state==False:
            st.markdown("---")
            st.subheader("六、計算成果")

            data = pd.DataFrame({
                'Q(cms）': [Q],
                'B（m）': [b],
                'n': [n],
                '1/s': [1/s],
                'd（m）': [y_ans],
                'A（m2）': [A],
                'P（m）': [P],
                'R': [R],
                'V（m/s）': [V],
                'hv（m）': [hv]
            })

            st.dataframe(data,hide_index=True)

            data = pd.DataFrame({
                '設計流量': [Q],
                '渠道寬度': [b],
                '曼寧係數': [n],
                '坡度': [1/s],
                '設計水深': [y_ans],
                '面積': [A],
                '潤周': [P],
                '水力半徑': [R],
                '設計流速': [V],
                '速度水頭': [hv]
            })

            st.dataframe(data,hide_index=True)

             #----Excel報表內容填寫----

        if (err_state==True and is_download==True) or err_state==False:

            workbook = openpyxl.load_workbook('./template/U2.xlsx')

            sheet = workbook.active

            # 将数字写入指定单元格，例如将数字 123 写入第一行第一列的单元格
            sheet.cell(row=1, column=1).value = st.session_state['company']
            sheet.cell(row=3, column=3).value = st.session_state['conName']
            sheet.cell(row=4, column=3).value = st.session_state['conLoc']
            
            sheet.cell(row=12, column=3).value = q
            sheet.cell(row=13, column=3).value = n
            sheet.cell(row=14, column=3).value = y
            sheet.cell(row=15, column=3).value = fill_angle
            sheet.cell(row=16, column=3).value = 1/s

            sheet.cell(row=21, column=3).value = b
            sheet.cell(row=22, column=3).value = Fb
            sheet.cell(row=24, column=8).value = P
            sheet.cell(row=25, column=5).value = A
            sheet.cell(row=26,column=5).value=R
            sheet.cell(row=27, column=6).value = V
            sheet.cell(row=28, column=5).value = Q

            sheet.cell(row=29, column=3).value = hv
            sheet.cell(row=30, column=3).value = Fr
            sheet.cell(row=32, column=3).value = Vc
            sheet.cell(row=33, column=3).value = yc_sol
            # sheet.cell(row=67, column=4).value = B_target

            output_file = 'example.xlsx'
            workbook.save(output_file)
            
            with open(output_file, 'rb') as f:
                bytes_data = f.read()
            st.download_button(label='計算成果下載', data=bytes_data, file_name=output_file,type='primary')

            os.remove(output_file)

    st.session_state.current_page = 'Uchannel_FixY'
